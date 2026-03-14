#!/usr/bin/env python3
"""
Generate RTMify OQ Fixture from the shipped template.

Usage: python3 gen_oq_fixture.py [--version <version>] [--out <path>]

Reads:  ../../site/public/RTMify_Requirements_Tracking_Template.xlsx
Writes: RTMify_OQ_Fixture_v<version>.xlsx by default

SCHEMA NOTES (discrepancies with prd.md):
  - User Needs tab has no "Derived Reqs" column. UN-OQ-005 (spec: dangling ref to
    REQ-OQ-999) becomes a second user_need_without_requirements advisory gap instead.
  - Tests tab has no "Linked Reqs" column. TST-OQ-004 (spec: dangling ref to REQ-OQ-999)
    becomes a test_group_without_requirements advisory gap instead.
  - The 4 E801 cross_ref diagnostics in expected-gaps.json were derived from those columns.
"""

from __future__ import annotations

import argparse
import os
import shutil
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
TEMPLATE = ROOT / "../../site/public/RTMify_Requirements_Tracking_Template.xlsx"
DEFAULT_VERSION = "0.1.0"


USER_NEEDS = [
    ("UN-OQ-001", "The system shall allow the operator to initiate a self-test sequence.", "Customer", "High"),
    ("UN-OQ-002", "The system shall log all operator actions with timestamps.", "Customer", "Medium"),
    ("UN-OQ-003", "The system shall support firmware updates in the field.", "Customer", "Medium"),
    ("UN-OQ-004", "The system shall detect and report sensor faults within 500ms.", "Customer", "High"),
    ("UN-OQ-005", "The system shall provide audible alarm for critical faults.", "Customer", "High"),
]

REQUIREMENTS = [
    (
        "REQ-OQ-001",
        "The software shall execute a power-on self-test (POST) and report pass/fail within 3 seconds of boot.",
        "UN-OQ-001",
        "High",
        "TST-OQ-001",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-002",
        "The software shall display POST results on the operator console within 1 second of test completion.",
        "UN-OQ-001",
        "Medium",
        "TST-OQ-002",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-003",
        "The software shall retain the last 1,000 operator log entries in non-volatile memory.",
        "UN-OQ-002",
        "Medium",
        None,
        "Approved",
        None,
    ),
    (
        "REQ-OQ-004",
        "The software shall enforce a minimum password length of 12 characters.",
        None,
        "Low",
        "TST-OQ-006",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-005",
        "The software shall detect sensor fault conditions within 500ms of occurrence.",
        "UN-OQ-004",
        "High",
        "TST-OQ-999",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-006",
        "Sensor fault reporting is expected to include the fault type and timestamp.",
        "UN-OQ-004",
        "Medium",
        "TST-OQ-003",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-007",
        "The software shall encrypt all data at rest using AES-256, and the software shall rotate encryption keys every 90 days.",
        "UN-OQ-002",
        "High",
        "TST-OQ-002",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-008",
        "The software shall write each operator action to the log within 200ms of the action.",
        "UN-OQ-002",
        "Medium",
        "TST-OQ-003",
        "Approved",
        None,
    ),
    (
        None,
        "The system shall support at least 50 concurrent users.",
        "UN-OQ-001",
        "Medium",
        "TST-OQ-001",
        "Approved",
        None,
    ),
    (
        "REQ-OQ-001",
        "The software shall provide a manual override for the self-test sequence.",
        "UN-OQ-001",
        "Low",
        "TST-OQ-001",
        "Approved",
        None,
    ),
]

TESTS = [
    ("TST-OQ-001", "T-OQ-001", "Verification", "Test", "Verify POST completes within 3 seconds of power-on and reports pass/fail. Status: Pass"),
    ("TST-OQ-002", "T-OQ-002", "Verification", "Test", "Verify POST results display on operator console within 1 second. Status: Pass"),
    ("TST-OQ-003", "T-OQ-003", "Verification", "Test", "Verify operator actions are logged within 200ms. Status: Fail"),
    ("TST-OQ-004", "T-OQ-004", "Verification", "Test", "Verify alarm volume exceeds 85dB at 1 meter. Status: Not Run"),
    ("TST-OQ-005", "T-OQ-005", "Verification", "Test", "Verify system survives 10,000 power cycles without data loss. Status: Pass"),
    ("TST-OQ-006", "T-OQ-006", "Verification", "Test", "Verify password enforcement rejects inputs under 12 characters. Status: Pass"),
]

RISKS = [
    (
        "RSK-OQ-001",
        "POST failure undetected, leading to operation with degraded sensor.",
        4,
        3,
        "POST runs automatically on every boot; failure blocks normal operation mode.",
        "REQ-OQ-001",
        1,
        1,
    ),
    (
        "RSK-OQ-002",
        "Sensor fault not detected within safety window, leading to incorrect output.",
        4,
        3,
        None,
        None,
        None,
        None,
    ),
    (
        "RSK-OQ-003",
        "Logging subsystem overflow causes loss of audit trail.",
        3,
        2,
        "Circular buffer overwrites oldest entries.",
        "REQ-OQ-003",
        4,
        2,
    ),
    (
        "RSK-OQ-004",
        "Unauthorized firmware update bricks field unit.",
        5,
        2,
        "Firmware update requires cryptographic signature verification.",
        "REQ-OQ-888",
        1,
        1,
    ),
]


def default_output_path(version: str) -> Path:
    return ROOT / f"RTMify_OQ_Fixture_v{version}.xlsx"


def clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def write_user_needs(ws) -> None:
    clear_data_rows(ws)
    for row_idx, (un_id, statement, source, priority) in enumerate(USER_NEEDS, start=2):
        ws.cell(row_idx, 1, un_id)
        ws.cell(row_idx, 2, statement)
        ws.cell(row_idx, 3, source)
        ws.cell(row_idx, 4, priority)


def write_requirements(ws) -> None:
    clear_data_rows(ws)
    for row_idx, (req_id, statement, un_id, priority, tg_ids, status, notes) in enumerate(REQUIREMENTS, start=2):
        ws.cell(row_idx, 1, req_id)
        ws.cell(row_idx, 2, statement)
        ws.cell(row_idx, 3, un_id)
        ws.cell(row_idx, 4, priority)
        ws.cell(row_idx, 5, tg_ids)
        ws.cell(row_idx, 6, status)
        ws.cell(row_idx, 7, notes)


def write_tests(ws) -> None:
    clear_data_rows(ws)
    for row_idx, (tg_id, t_id, t_type, method, notes) in enumerate(TESTS, start=2):
        ws.cell(row_idx, 1, tg_id)
        ws.cell(row_idx, 2, t_id)
        ws.cell(row_idx, 3, t_type)
        ws.cell(row_idx, 4, method)
        ws.cell(row_idx, 5, notes)


def write_risks(ws) -> None:
    clear_data_rows(ws)
    for row_idx, (risk_id, desc, init_sev, init_like, mitigation, linked_req, res_sev, res_like) in enumerate(RISKS, start=2):
        ws.cell(row_idx, 1, risk_id)
        ws.cell(row_idx, 2, desc)
        ws.cell(row_idx, 3, init_sev)
        ws.cell(row_idx, 4, init_like)
        ws.cell(row_idx, 5, mitigation)
        ws.cell(row_idx, 6, linked_req)
        ws.cell(row_idx, 7, res_sev)
        ws.cell(row_idx, 8, res_like)


def build_fixture(output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE, output_path)
    wb = openpyxl.load_workbook(output_path)
    write_user_needs(wb["User Needs"])
    write_requirements(wb["Requirements"])
    write_tests(wb["Tests"])
    write_risks(wb["Risks"])
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate the RTMify OQ validation fixture.")
    parser.add_argument("--version", default=DEFAULT_VERSION, help="Version suffix for the default output filename.")
    parser.add_argument("--out", type=Path, help="Explicit output path for the generated workbook.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.out or default_output_path(args.version)
    build_fixture(output)
    print(f"Written: {output}")
    print()
    print("Expected output from rtmify-trace:")
    print("  Hard gaps (gap_count):   5")
    print("  Advisory gaps:           4  (UN-OQ-003, UN-OQ-005, TST-OQ-004, TST-OQ-005)")
    print("  Diagnostic warnings:     7")
    print("  Diagnostic errors:       0")
    print()
    print("NOTE: UN-OQ-005 and TST-OQ-004 produce advisory gaps")
    print("because the template has no Derived Reqs / Linked Reqs columns.")


if __name__ == "__main__":
    main()
